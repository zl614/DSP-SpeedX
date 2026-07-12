#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SpeedX 操作员人效面板  — 自带登录 · 单文件本地程序
=====================================================
双击运行后会：① 起一个本地服务 ② 自动打开浏览器到登录页。
每个人用【自己的】SpeedX 账号登录（账号密码只在本机内存，不保存文件、不上传）。
登录后即可：选批次 → 拉数据 → 看面板 → 导出 CSV。

整条链路（都在本机完成，数据不出本机）：
    登录 → 找批次 → 取/生成报告 → 读 "Parcel Total Scan Result" 表 → 出 CSV → 面板渲染

本地直接跑（开发/Mac 验证）:
    pip install flask requests openpyxl          # 若报 externally-managed 加 --break-system-packages
    # 可选(更快读大表): pip install python-calamine
    python3 speedx_relay.py                       # 浏览器自动打开 http://127.0.0.1:5057/
打 Windows exe：见同目录 BUILD_GITHUB.md（GitHub 云端构建）。

注意:本程序需与 dashboard_realtime.html、exception_reason_analysis.html 放在一起（打包时已一并打入 exe）。
      启动后默认打开操作员人效面板(/)；异常原因分析在 /exceptions（两页顶部有互相跳转的链接）。
"""
import io
import csv
import time
import json
import os
import sys
import tempfile
import datetime
import re
import threading
import webbrowser
from pathlib import Path

from flask import Flask, request, Response
import requests

# ===================== 配置 =====================
CFG = {
    "base_url": "https://pos.speedx.io",
    "warehouse": "EWR",         # 默认仓（仅影响"留空=当前ACTIVE"的批次匹配；选批次时不受影响）
    "date_window_days": 14,
    "poll_timeout_sec": 300,
    "poll_interval_sec": 5,
    "api_timeout_sec": 90,      # 单次接口超时（SpeedX 偶尔慢）
    "port": 5057,
}
BASE = CFG["base_url"].rstrip("/")

# 登录凭据：由网页登录页填入，只存在内存中，绝不写文件、不外传。
CREDS = {"login_name": None, "password": None, "facility": None, "verify_code": "8888"}
LOGGED_IN = {"ok": False}

LOGIN_PATH        = "/pos-exchange-api/login/login"
PARENT_LIST_PATH  = "/pos-exchange-api/web/inbound/batch/parentList"
PARENT_REPORT_PATH= "/pos-exchange-api/web/inbound/batch/parentReport/v2"
DC_LIST_PATH      = "/pos-exchange-api/web/report-task/list"
TASK_TYPE         = "INBOUND_BATCH_PARENT_REPORT"
SUCCESS_WORDS     = ("SUCCESS", "SUCCEED", "SUCCEEDED", "DONE", "FINISHED", "COMPLETED")

SHEET_NAME = "Parcel Total Scan Result"
# 面板需要的列（输出顺序）。源表头名与这些完全一致。
OUT_COLS = ["Tracking Number", "Scan Operator", "Scan Time", "Status", "Facility Code", "Scan Facility Code"]

# ===================== Outlook 邮件（Microsoft Graph）=====================
# 新版 Outlook 没有 COM 接口，本地脚本驱动不了它；能发信的只剩 Graph API。
# 这里全部用 requests 手写 OAuth（设备码流程），不引入新依赖，spec 不用改。
#
# 配置方式（三选一，优先级从高到低）：
#   1) 环境变量 SPEEDX_GRAPH_CLIENT_ID / SPEEDX_GRAPH_TENANT_ID
#   2) 与本程序（或 exe）同目录的 graph_config.json：
#        {"client_id": "xxx", "tenant_id": "yyy", "allow_draft": true}
#   3) 下面 GRAPH_CFG 里的默认值
# 没配 client_id 也不影响：拆分器照常能用，只是"发送"按钮是灰的。
GRAPH_CFG = {
    "client_id": "",          # ← Azure 应用注册的 Application (client) ID
    "tenant_id": "common",    # ← Directory (tenant) ID（单租户应用必须填真实 tenant id 或 speedx.io）
    "allow_draft": True,      # True=可建草稿(需 Mail.ReadWrite)；False=只发送(仅需 Mail.Send)
}
GRAPH_LOGIN = "https://login.microsoftonline.com"
GRAPH_API   = "https://graph.microsoft.com/v1.0"

GRAPH_TOKEN = {"access": None, "refresh": None, "exp": 0.0, "user": ""}
GRAPH_FLOW  = {"device_code": None, "interval": 5, "expires": 0.0}

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _graph_scopes():
    s = ["https://graph.microsoft.com/Mail.Send", "https://graph.microsoft.com/User.Read", "offline_access"]
    if GRAPH_CFG["allow_draft"]:
        s.insert(1, "https://graph.microsoft.com/Mail.ReadWrite")
    return " ".join(s)


def _authority():
    return f"{GRAPH_LOGIN}/{(GRAPH_CFG['tenant_id'] or 'common').strip()}"


def _app_dir():
    """exe 时取 exe 所在目录（改配置不用重新打包）；开发时取脚本目录。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def load_graph_cfg():
    cid = os.environ.get("SPEEDX_GRAPH_CLIENT_ID")
    tid = os.environ.get("SPEEDX_GRAPH_TENANT_ID")
    if cid:
        GRAPH_CFG["client_id"] = cid.strip()
    if tid:
        GRAPH_CFG["tenant_id"] = tid.strip()
    p = os.path.join(_app_dir(), "graph_config.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                j = json.load(f)
            if j.get("client_id"):
                GRAPH_CFG["client_id"] = str(j["client_id"]).strip()
            if j.get("tenant_id"):
                GRAPH_CFG["tenant_id"] = str(j["tenant_id"]).strip()
            if "allow_draft" in j:
                GRAPH_CFG["allow_draft"] = bool(j["allow_draft"])
        except Exception as e:
            print(f"[mail] graph_config.json 读取失败：{e}")
    return bool(GRAPH_CFG["client_id"])


def _store_token(j):
    GRAPH_TOKEN["access"] = j.get("access_token")
    GRAPH_TOKEN["refresh"] = j.get("refresh_token") or GRAPH_TOKEN["refresh"]
    GRAPH_TOKEN["exp"] = time.time() + int(j.get("expires_in", 3600)) - 120
    try:
        me = requests.get(f"{GRAPH_API}/me",
                          headers={"Authorization": "Bearer " + GRAPH_TOKEN["access"]},
                          timeout=20).json()
        GRAPH_TOKEN["user"] = me.get("mail") or me.get("userPrincipalName") or ""
    except Exception:
        pass


def graph_token():
    """拿可用的 access token；过期就用 refresh token 续。"""
    if GRAPH_TOKEN["access"] and time.time() < GRAPH_TOKEN["exp"]:
        return GRAPH_TOKEN["access"]
    if GRAPH_TOKEN["refresh"]:
        r = requests.post(f"{_authority()}/oauth2/v2.0/token", data={
            "grant_type": "refresh_token",
            "client_id": GRAPH_CFG["client_id"],
            "refresh_token": GRAPH_TOKEN["refresh"],
            "scope": _graph_scopes(),
        }, timeout=30)
        j = r.json()
        if "access_token" in j:
            _store_token(j)
            return GRAPH_TOKEN["access"]
    raise RuntimeError("邮箱未登录或登录已过期，请重新登录。")


def _recips(s):
    return [{"emailAddress": {"address": a.strip()}}
            for a in re.split(r"[;,]", s or "") if a.strip()]


SESSION = requests.Session()
SESSION.headers.update({"Accept": "application/json", "Content-Type": "application/json"})


# ===================== 登录 + 请求 =====================
def login():
    if not CREDS["password"]:
        raise RuntimeError("尚未登录")
    r = SESSION.post(BASE + LOGIN_PATH, json={
        "loginName": CREDS["login_name"], "passwd": CREDS["password"], "autoLogin": True,
        "facilityCode": CREDS["facility"], "verifyCode": str(CREDS["verify_code"] or "8888"),
    }, timeout=CFG.get("api_timeout_sec", 90))
    r.raise_for_status()
    d = r.json()
    tok = (d.get("payload") or {}).get("authToken")
    if not (d.get("success") and tok):
        raise RuntimeError(f"登录失败: {d.get('message')}")
    SESSION.headers["Authorization"] = tok
    LOGGED_IN["ok"] = True


def api_post(path, payload, _retry=True, _timeout_retry=True):
    try:
        r = SESSION.post(BASE + path, json=payload, timeout=CFG.get("api_timeout_sec", 90))
    except requests.exceptions.Timeout:
        if _timeout_retry:
            time.sleep(2)
            return api_post(path, payload, _retry=_retry, _timeout_retry=False)
        raise
    data = None
    need = r.status_code in (401, 403)
    if not need and r.ok:
        data = r.json()
        if isinstance(data, dict) and data.get("success") is False:
            m = str(data.get("message", "")).lower()
            if any(x in m for x in ("unauthorized", "token", "登录", "login")):
                need = True
    if need and _retry:
        login()
        return api_post(path, payload, _retry=False, _timeout_retry=_timeout_retry)
    r.raise_for_status()
    return data if data is not None else r.json()


def _first(d, keys, dv=None):
    for k in keys:
        v = d.get(k)
        if v not in (None, ""):
            return v
    return dv


def _rows(resp):
    if isinstance(resp, dict):
        for w in ("payload", "data", "result"):
            inner = resp.get(w)
            if isinstance(inner, (dict, list)):
                resp = inner
                break
    if isinstance(resp, list):
        return resp
    if isinstance(resp, dict):
        for k in ("list", "records", "rows", "content", "items"):
            if isinstance(resp.get(k), list):
                return resp[k]
    return []


# ===================== 业务：批次 / 报告 =====================
def find_active_batch():
    now = int(time.time() * 1000)
    win = CFG["date_window_days"] * 86400 * 1000
    rows = _rows(api_post(PARENT_LIST_PATH, {
        "status": "ACTIVE", "testFlag": False,
        "createDateStart": now - win, "createDateEnd": now + 86400000,
        "order": "desc", "sidx": "id", "pageIndex": 0, "pageSize": 20,
    }))
    if not rows:
        raise RuntimeError("没查到 ACTIVE 父批次。")
    pat = re.compile(r"^\d{6}-P-" + re.escape(CFG["warehouse"]) + r"-\d+$")
    real = [b for b in rows if pat.match(str(_first(b, ["batchCode", "parentBatchCode", "code"], "")))]
    b = (real or rows)[0]
    return str(_first(b, ["id", "batchId"])), _first(b, ["batchCode", "parentBatchCode", "code"])


def _date_to_ms(date_str, end_of_day=False):
    y, m, d = map(int, str(date_str).split("-"))
    dt = (datetime.datetime(y, m, d, 23, 59, 59, 999000) if end_of_day
          else datetime.datetime(y, m, d, 0, 0, 0))
    return int(dt.timestamp() * 1000)


def list_batches(status="ALL", start_ms=None, end_ms=None):
    """列出可选父批次（供面板下拉）。status='ALL' 则不按状态过滤。"""
    now = int(time.time() * 1000)
    if start_ms is None:
        start_ms = now - CFG["date_window_days"] * 86400 * 1000
    if end_ms is None:
        end_ms = now + 86400000
    payload = {"testFlag": False, "createDateStart": start_ms, "createDateEnd": end_ms,
               "order": "desc", "sidx": "id", "pageIndex": 0, "pageSize": 30}
    if status and str(status).upper() != "ALL":
        payload["status"] = status
    out = []
    for b in _rows(api_post(PARENT_LIST_PATH, payload)):
        out.append({
            "batchCode": _first(b, ["batchCode", "parentBatchCode", "code"]),
            "id": _first(b, ["id", "batchId"]),
            "routeDate": _first(b, ["fixRouteDate", "routeDate"]),
            "status": _first(b, ["status"]),
            "totalQty": _first(b, ["parcelTotalQty", "totalQty", "parcelHubQty"]),
        })
    return out


def find_batch_by_code(code, status="ALL"):
    """按批次号找 id（先近 14 天，找不到再放宽到近 60 天）。"""
    for window_days in (CFG["date_window_days"], 60):
        now = int(time.time() * 1000)
        for b in list_batches(status, now - window_days * 86400 * 1000, now + 86400000):
            if str(b["batchCode"]) == str(code):
                return str(b["id"]), b["batchCode"]
    raise RuntimeError(f"没找到批次 {code}（可能不在近 60 天，或状态过滤把它排除了）。")


def _dc_payload(page_size=50):
    now = int(time.time() * 1000)
    days = CFG["date_window_days"]
    return {"businessType": TASK_TYPE, "createDateStart": now - days * 86400 * 1000,
            "createDateEnd": now + 86400000, "order": "desc", "sidx": "id",
            "pageIndex": 0, "pageNum": 1, "pageSize": page_size}


def latest_report_url(parent_code):
    """当前批次最近一份已 SUCCESS 的报告文件 URL（没有则 None）。"""
    for t in _rows(api_post(DC_LIST_PATH, _dc_payload())):
        st = str(_first(t, ["status", "taskStatus", "state"], "")).upper()
        url = _first(t, ["businessResult", "fileUrl", "url", "downloadUrl", "filePath"], "")
        if st in SUCCESS_WORDS and url and str(parent_code) in json.dumps(t, ensure_ascii=False):
            return url
    return None


def trigger_and_wait(parent_id, parent_code):
    known = {_first(t, ["id"]) for t in _rows(api_post(DC_LIST_PATH, _dc_payload()))}
    api_post(PARENT_REPORT_PATH, {"businessData": str(parent_id), "businessNo": str(int(time.time() * 1000))})
    deadline = time.time() + CFG["poll_timeout_sec"]
    while time.time() < deadline:
        for t in _rows(api_post(DC_LIST_PATH, _dc_payload())):
            if _first(t, ["id"]) in known:
                continue
            st = str(_first(t, ["status", "taskStatus", "state"], "")).upper()
            url = _first(t, ["businessResult", "fileUrl", "url", "downloadUrl", "filePath"], "")
            if st in SUCCESS_WORDS and url and str(parent_code) in json.dumps(t, ensure_ascii=False):
                return url
            if st in ("FAILED", "DISCARD"):
                raise RuntimeError(f"报告任务失败: {st}")
        time.sleep(CFG["poll_interval_sec"])
    raise TimeoutError("等待报告生成超时。")


def download_xlsx(url):
    """SAS 链接，直接 GET（不带 Authorization）。存临时文件，返回路径。"""
    tmp = Path(tempfile.gettempdir()) / f"speedx_scan_{int(time.time())}.xlsx"
    with requests.get(url, stream=True, timeout=300) as r:
        r.raise_for_status()
        with open(tmp, "wb") as f:
            for chunk in r.iter_content(1 << 16):
                if chunk:
                    f.write(chunk)
    return tmp


# ===================== 读 xlsx 的目标表 → CSV =====================
def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        # 面板按 MM/DD/YYYY HH:MM:SS 解析（取前2位月、4-5位日、12-13位时）
        try:
            return v.strftime("%m/%d/%Y %H:%M:%S")
        except Exception:
            return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _find_header(matrix_iter_first_rows):
    """在前几行里找真正的表头行（含 Scan Operator + Scan Time），返回 (header_list, header_row_index)。"""
    for i, row in enumerate(matrix_iter_first_rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Scan Operator" in cells and "Scan Time" in cells:
            return cells, i
    return None, -1


def scan_sheet_to_csv(xlsx_path):
    """优先用 calamine（快、且无视错误的 dimension），否则 openpyxl（read_only + 重置维度）。返回 CSV 文本。"""
    # ---- 尝试 calamine ----
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(str(xlsx_path))
        data = wb.get_sheet_by_name(SHEET_NAME).to_python()  # list[list]
        header, hidx = _find_header(data[:8])
        if not header:
            raise RuntimeError(f"在表「{SHEET_NAME}」前几行没找到表头（Scan Operator/Scan Time）。")
        idx = {c: header.index(c) for c in OUT_COLS if c in header}
        missing = [c for c in OUT_COLS if c not in idx]
        if missing:
            raise RuntimeError(f"表里缺列: {missing}（实际表头: {header}）")
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(OUT_COLS)
        n = 0
        for row in data[hidx + 1:]:
            if not row:
                continue
            w.writerow([_fmt(row[idx[c]]) if idx[c] < len(row) else "" for c in OUT_COLS])
            n += 1
        print(f"[读表] calamine 读到 {n} 行数据")
        return buf.getvalue()
    except ImportError:
        pass  # 没装 calamine，走 openpyxl

    # ---- openpyxl read_only ----
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"报告里没有「{SHEET_NAME}」表，实际有: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    # 关键：报告 xlsx 内部声明的数据范围(dimension)可能是错的（只声明 1 行），
    # read_only 模式会据此截断。重置维度后强制按实际单元格读全部行。
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    it = ws.iter_rows(values_only=True)
    first8 = []
    for _ in range(8):
        try:
            first8.append(next(it))
        except StopIteration:
            break
    header, hidx = _find_header(first8)
    if not header:
        raise RuntimeError(f"在表「{SHEET_NAME}」前几行没找到表头（Scan Operator/Scan Time）。")
    idx = {c: header.index(c) for c in OUT_COLS if c in header}
    missing = [c for c in OUT_COLS if c not in idx]
    if missing:
        raise RuntimeError(f"表里缺列: {missing}（实际表头: {header}）")
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(OUT_COLS)
    n = 0
    for row in first8[hidx + 1:]:          # first8 里 header 之后的几行也要写
        if row and any(v is not None for v in row):
            w.writerow([_fmt(row[idx[c]]) if idx[c] < len(row) else "" for c in OUT_COLS])
            n += 1
    for row in it:
        if row and any(v is not None for v in row):
            w.writerow([_fmt(row[idx[c]]) if idx[c] < len(row) else "" for c in OUT_COLS])
            n += 1
    wb.close()
    print(f"[读表] openpyxl 读到 {n} 行数据")
    return buf.getvalue()


# ===================== Flask =====================
app = Flask(__name__)


def resource_path(rel):
    """兼容 PyInstaller 打包：打包后从 _MEIPASS 读，开发时从脚本目录读。"""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def find_asset(rel):
    """先找打包进 exe 里的（_MEIPASS / 脚本目录），找不到再找 exe 同目录。
    这样新页面即使没写进 .spec，把 html 放到 exe 旁边一样能用；
    以后改页面也不用重新打包——换掉 exe 旁边那个 html 就行。"""
    p = resource_path(rel)
    if os.path.exists(p):
        return p
    q = os.path.join(_app_dir(), rel)
    return q if os.path.exists(q) else p


PAGE_FILE = resource_path("dashboard_realtime.html")
EXC_FILE = resource_path("exception_reason_analysis.html")
# fleet_splitter.html 在请求时用 find_asset() 找（_MEIPASS 或 exe 同目录都行）


@app.after_request
def cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return resp


@app.route("/")
def index():
    try:
        with open(PAGE_FILE, encoding="utf-8") as f:
            return Response(f.read(), mimetype="text/html; charset=utf-8")
    except FileNotFoundError:
        return Response("找不到 dashboard_realtime.html（应与程序在一起）。", status=500,
                        mimetype="text/plain; charset=utf-8")


@app.route("/exceptions")
def exceptions_page():
    try:
        with open(EXC_FILE, encoding="utf-8") as f:
            return Response(f.read(), mimetype="text/html; charset=utf-8")
    except FileNotFoundError:
        return Response("找不到 exception_reason_analysis.html（应与程序在一起）。", status=500,
                        mimetype="text/plain; charset=utf-8")


@app.route("/split")
def split_page():
    path = find_asset("fleet_splitter.html")
    try:
        with open(path, encoding="utf-8") as f:
            return Response(f.read(), mimetype="text/html; charset=utf-8")
    except FileNotFoundError:
        return Response(
            "找不到 fleet_splitter.html。\n"
            "把它放到 exe 同目录即可（或写进 .spec 的 datas 一起打包）。",
            status=500, mimetype="text/plain; charset=utf-8")


@app.route("/login", methods=["POST"])
def do_login():
    data = request.get_json(force=True, silent=True) or {}
    CREDS["login_name"] = (data.get("loginName") or "").strip()
    CREDS["password"] = data.get("passwd") or ""
    CREDS["facility"] = (data.get("facilityCode") or "").strip()
    CREDS["verify_code"] = str(data.get("verifyCode") or "8888").strip()
    if not CREDS["login_name"] or not CREDS["password"] or not CREDS["facility"]:
        CREDS["password"] = ""
        return {"ok": False, "msg": "用户名 / 密码 / Facility 都要填。"}
    try:
        login()
        return {"ok": True, "loginName": CREDS["login_name"]}
    except Exception as e:
        LOGGED_IN["ok"] = False
        CREDS["password"] = ""                  # 失败立即清掉密码
        return {"ok": False, "msg": str(e)}


@app.route("/logout", methods=["POST"])
def do_logout():
    CREDS["password"] = ""
    LOGGED_IN["ok"] = False
    SESSION.headers.pop("Authorization", None)
    return {"ok": True}


@app.route("/me")
def me():
    return {"loggedIn": LOGGED_IN["ok"], "loginName": CREDS["login_name"] or ""}


@app.route("/batches")
def batches():
    status = request.args.get("status", "ALL")
    start = request.args.get("start")
    end = request.args.get("end")
    try:
        login()
        bs = list_batches(status,
                          _date_to_ms(start, False) if start else None,
                          _date_to_ms(end, True) if end else None)
        return {"ok": True, "batches": bs}
    except Exception as e:
        return {"ok": False, "msg": str(e)}, 500


@app.route("/scan-csv")
def scan_csv():
    fresh = request.args.get("fresh", "0") in ("1", "true", "yes")
    batch = request.args.get("batch")            # 批次号（显示用）
    batch_id = request.args.get("batchId")       # 批次内部 id（面板从 /batches 直接带来，最快、免超时）
    status = request.args.get("status", "ALL")
    try:
        login()
        if batch_id:                              # 有 id 直接用，跳过二次查询（也跳过超时）
            parent_id, parent_code = str(batch_id), (batch or str(batch_id))
        elif batch:
            parent_id, parent_code = find_batch_by_code(batch, status)
        else:
            parent_id, parent_code = find_active_batch()
        if fresh:
            url = trigger_and_wait(parent_id, parent_code)
        else:
            url = latest_report_url(parent_code) or trigger_and_wait(parent_id, parent_code)
        xlsx_path = download_xlsx(url)
        try:
            csv_text = scan_sheet_to_csv(xlsx_path)
        finally:
            try:
                xlsx_path.unlink()
            except Exception:
                pass
        return Response(csv_text, mimetype="text/csv; charset=utf-8",
                        headers={"X-Parent-Batch": str(parent_code)})
    except Exception as e:
        import traceback
        traceback.print_exc()                      # 完整堆栈打到终端
        return Response(f"ERROR: {type(e).__name__}: {e}", status=500,
                        mimetype="text/plain; charset=utf-8")


# ---------- 邮件：登录 / 状态 / 发送 ----------
@app.route("/mail/status")
def mail_status():
    return {
        "configured": bool(GRAPH_CFG["client_id"]),
        "signedIn": bool(GRAPH_TOKEN["access"] or GRAPH_TOKEN["refresh"]),
        "user": GRAPH_TOKEN["user"] or "",
        "canDraft": bool(GRAPH_CFG["allow_draft"]),
    }


@app.route("/mail/signin-start", methods=["POST"])
def mail_signin_start():
    if not GRAPH_CFG["client_id"]:
        return {"ok": False, "msg": "还没配置 client_id（把 graph_config.json 放到程序同目录）。"}
    try:
        r = requests.post(f"{_authority()}/oauth2/v2.0/devicecode",
                          data={"client_id": GRAPH_CFG["client_id"], "scope": _graph_scopes()},
                          timeout=30)
        j = r.json()
    except Exception as e:
        return {"ok": False, "msg": f"连不上 Microsoft 登录服务：{e}"}
    if "device_code" not in j:
        return {"ok": False, "msg": j.get("error_description") or str(j)[:300]}
    GRAPH_FLOW["device_code"] = j["device_code"]
    GRAPH_FLOW["interval"] = int(j.get("interval", 5))
    GRAPH_FLOW["expires"] = time.time() + int(j.get("expires_in", 900))
    return {"ok": True,
            "userCode": j.get("user_code"),
            "verificationUri": j.get("verification_uri") or "https://microsoft.com/devicelogin",
            "interval": GRAPH_FLOW["interval"]}


@app.route("/mail/signin-poll", methods=["POST"])
def mail_signin_poll():
    if not GRAPH_FLOW["device_code"]:
        return {"ok": False, "state": "none", "msg": "还没开始登录。"}
    if time.time() > GRAPH_FLOW["expires"]:
        GRAPH_FLOW["device_code"] = None
        return {"ok": False, "state": "expired", "msg": "登录码过期了，重新点一次登录。"}
    try:
        r = requests.post(f"{_authority()}/oauth2/v2.0/token", data={
            "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
            "client_id": GRAPH_CFG["client_id"],
            "device_code": GRAPH_FLOW["device_code"],
        }, timeout=30)
        j = r.json()
    except Exception as e:
        return {"ok": False, "state": "error", "msg": str(e)}
    if "access_token" in j:
        GRAPH_FLOW["device_code"] = None
        _store_token(j)
        return {"ok": True, "state": "done", "user": GRAPH_TOKEN["user"]}
    err = j.get("error", "")
    if err in ("authorization_pending", "slow_down"):
        return {"ok": True, "state": "pending"}
    GRAPH_FLOW["device_code"] = None
    return {"ok": False, "state": "error",
            "msg": j.get("error_description") or err or "登录失败"}


@app.route("/mail/signout", methods=["POST"])
def mail_signout():
    GRAPH_TOKEN.update({"access": None, "refresh": None, "exp": 0.0, "user": ""})
    GRAPH_FLOW["device_code"] = None
    return {"ok": True}


@app.route("/mail/send-one", methods=["POST"])
def mail_send_one():
    """发一封（或建一封草稿）。前端循环调用，好显示进度、好定位是哪封失败。"""
    d = request.get_json(force=True, silent=True) or {}
    mode = (d.get("mode") or "draft").lower()
    to = (d.get("to") or "").strip()
    if not to:
        return {"ok": False, "msg": "没有收件人"}
    if mode == "draft" and not GRAPH_CFG["allow_draft"]:
        return {"ok": False, "msg": "草稿功能没开（需要 Mail.ReadWrite 权限）"}

    file_b64 = d.get("fileB64") or ""
    file_name = d.get("fileName") or "attachment.xlsx"
    if len(file_b64) > 3_000_000:                      # Graph 内联附件上限 ~3MB
        return {"ok": False, "msg": f"{file_name} 超过 3MB，Graph 内联附件放不下"}

    msg = {
        "subject": d.get("subject") or "",
        "body": {"contentType": "Text", "content": d.get("body") or ""},
        "toRecipients": _recips(to),
        "attachments": [{
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": file_name,
            "contentType": XLSX_MIME,
            "contentBytes": file_b64,
        }],
    }
    cc = (d.get("cc") or "").strip()
    if cc:
        msg["ccRecipients"] = _recips(cc)

    try:
        tok = graph_token()
    except Exception as e:
        return {"ok": False, "needLogin": True, "msg": str(e)}

    h = {"Authorization": "Bearer " + tok, "Content-Type": "application/json"}
    try:
        if mode == "send":
            r = requests.post(f"{GRAPH_API}/me/sendMail", headers=h,
                              json={"message": msg, "saveToSentItems": True}, timeout=120)
            ok = r.status_code in (200, 202)
        else:
            r = requests.post(f"{GRAPH_API}/me/messages", headers=h, json=msg, timeout=120)
            ok = r.status_code in (200, 201)
        if ok:
            return {"ok": True}
        return {"ok": False, "msg": f"Graph {r.status_code}: {r.text[:300]}"}
    except Exception as e:
        return {"ok": False, "msg": f"{type(e).__name__}: {e}"}


@app.route("/health")
def health():
    return {"ok": True}


def _open_browser(port):
    webbrowser.open(f"http://127.0.0.1:{port}/")


if __name__ == "__main__":
    port = CFG["port"]
    ok_graph = load_graph_cfg()
    print(f"SpeedX 面板已启动 → 浏览器打开 http://127.0.0.1:{port}/ （登录你自己的账号）")
    print(f"  拆分器 / 发邮件 → http://127.0.0.1:{port}/split")
    if ok_graph:
        print(f"  [mail] Graph 已配置 (client_id …{GRAPH_CFG['client_id'][-6:]}, tenant={GRAPH_CFG['tenant_id']})")
    else:
        print("  [mail] 未配置 Graph client_id → 拆分照常用，但『发送』是灰的。")
        print("         把 graph_config.json 放到本程序同目录即可启用。")
    # 仅主进程开浏览器（避免 reloader 重复；这里 debug=False 本就单进程）
    threading.Timer(1.2, _open_browser, args=(port,)).start()
    app.run(host="127.0.0.1", port=port, debug=False)
